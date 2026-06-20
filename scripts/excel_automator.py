"""
Excel Automator - Bulk operations on Excel files
Usage: python excel_automator.py merge --input "*.xlsx" --output merged.xlsx
       python excel_automator.py clean --input data.xlsx --output cleaned.xlsx
       python excel_automator.py pivot --input data.xlsx --index "Category" --values "Sales"

Features:
- Merge multiple Excel/CSV files
- Clean & deduplicate data
- Create pivot tables & charts
- Auto-format with colors & borders
- Generate summary statistics
"""

import pandas as pd
import glob
import argparse
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

def merge_files(input_pattern, output_file):
    """Merge multiple CSV/Excel files into one."""
    files = glob.glob(input_pattern)
    if not files:
        print(f"No files found matching: {input_pattern}")
        return
    
    dfs = []
    for f in files:
        ext = os.path.splitext(f)[1].lower()
        if ext == '.csv':
            df = pd.read_csv(f)
        else:
            df = pd.read_excel(f)
        df['_source_file'] = os.path.basename(f)
        dfs.append(df)
        print(f"  Loaded: {f} ({len(df)} rows)")
    
    merged = pd.concat(dfs, ignore_index=True)
    
    ext = os.path.splitext(output_file)[1].lower()
    if ext == '.csv':
        merged.to_csv(output_file, index=False)
    else:
        merged.to_excel(output_file, index=False)
    
    print(f"\nMerged {len(files)} files → {len(merged)} rows → {output_file}")

def clean_data(input_file, output_file):
    """Clean and deduplicate Excel/CSV data."""
    ext = os.path.splitext(input_file)[1].lower()
    df = pd.read_csv(input_file) if ext == '.csv' else pd.read_excel(input_file)
    
    original_rows = len(df)
    
    # Remove duplicates
    df = df.drop_duplicates()
    
    # Remove empty rows/columns
    df = df.dropna(how='all')
    df = df.dropna(axis=1, how='all')
    
    # Strip whitespace from string columns
    str_cols = df.select_dtypes(include=['object']).columns
    df[str_cols] = df[str_cols].apply(lambda x: x.str.strip())
    
    # Standardize column names
    df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')
    
    df.to_excel(output_file, index=False)
    
    removed = original_rows - len(df)
    print(f"Cleaned: {original_rows} → {len(df)} rows ({removed} removed)")
    print(f"Output: {output_file}")

def create_pivot(input_file, output_file, index_col, values_col, aggfunc='sum'):
    """Create a pivot table with auto-chart."""
    ext = os.path.splitext(input_file)[1].lower()
    df = pd.read_csv(input_file) if ext == '.csv' else pd.read_excel(input_file)
    
    pivot = pd.pivot_table(df, index=index_col, values=values_col, aggfunc=aggfunc)
    pivot = pivot.sort_values(values_col, ascending=False)
    
    # Write to Excel with chart
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        pivot.to_excel(writer, sheet_name='Pivot')
        df.to_excel(writer, sheet_name='Raw Data', index=False)
    
    # Add chart
    wb = load_workbook(output_file)
    ws = wb['Pivot']
    
    chart = BarChart()
    chart.title = f"{aggfunc.upper()} of {values_col} by {index_col}"
    chart.y_axis.title = values_col
    chart.x_axis.title = index_col
    
    data_ref = Reference(ws, min_col=2, min_row=1, max_row=len(pivot)+1)
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=len(pivot)+1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.shape = 4
    ws.add_chart(chart, "D2")
    
    wb.save(output_file)
    print(f"Pivot table created → {output_file}")

def format_excel(input_file, output_file=None):
    """Auto-format Excel with colors, borders, and auto-width."""
    if output_file is None:
        output_file = input_file
    
    wb = load_workbook(input_file)
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for ws in wb.worksheets:
        # Format headers
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Auto-width
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 3, 50)
        
        # Add borders
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
        
        # Zebra striping
        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if i % 2 == 0:
                for cell in row:
                    cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    
    wb.save(output_file)
    print(f"Formatted: {output_file}")

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Excel Automator')
    sub = parser.add_subparsers(dest='command')
    
    p_merge = sub.add_parser('merge', help='Merge multiple files')
    p_merge.add_argument('--input', required=True, help='File pattern (e.g., "*.xlsx")')
    p_merge.add_argument('--output', required=True, help='Output file')
    
    p_clean = sub.add_parser('clean', help='Clean and deduplicate')
    p_clean.add_argument('--input', required=True)
    p_clean.add_argument('--output', required=True)
    
    p_pivot = sub.add_parser('pivot', help='Create pivot table')
    p_pivot.add_argument('--input', required=True)
    p_pivot.add_argument('--output', required=True)
    p_pivot.add_argument('--index', required=True, help='Column for rows')
    p_pivot.add_argument('--values', required=True, help='Column for values')
    p_pivot.add_argument('--aggfunc', default='sum', choices=['sum', 'mean', 'count'])
    
    p_fmt = sub.add_parser('format', help='Auto-format Excel')
    p_fmt.add_argument('--input', required=True)
    p_fmt.add_argument('--output')
    
    args = parser.parse_args()
    
    if args.command == 'merge':
        merge_files(args.input, args.output)
    elif args.command == 'clean':
        clean_data(args.input, args.output)
    elif args.command == 'pivot':
        create_pivot(args.input, args.output, args.index, args.values, args.aggfunc)
    elif args.command == 'format':
        format_excel(args.input, args.output)
    else:
        parser.print_help()
